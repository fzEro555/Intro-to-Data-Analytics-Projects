# -*- coding: utf-8 -*-
"""
Created on Fri Oct  5 09:34:12 2018

@author: Subha Karanam
"""

import numpy as np
import pandas as pd
import openpyxl
import pprint

from OneDataFrame import *
from Cleaning import *

def main() :
    #removing summary page 
    wb = openpyxl.load_workbook('Hurricane_Maria_Report.xlsx')
    workbook_Maria = OneDataFrame.removeSummary(wb)
    workbook_Maria.save('Hurricane_Maria_Report_new.xlsx')
    wb = openpyxl.load_workbook('Hurricane_Harvey_Report.xlsx')
    workbook_Harvey = OneDataFrame.removeSummary(wb)
    workbook_Harvey.save('Hurricane_Harvey_Report_new.xlsx')
    wb = openpyxl.load_workbook('Hurricane_Irma_Report.xlsx')
    workbook_Irma = OneDataFrame.removeSummary(wb)
    workbook_Harvey.save('Hurricane_Irma_Report_new.xlsx')
    wb = openpyxl.load_workbook('Hurricane_Irene_Report.xlsx')
    workbook_Maria = OneDataFrame.removeSummary(wb)
    workbook_Maria.save('Hurricane_Irene_Report_new.xlsx')

    #getting a pandas dataframe without headers
    df_Maria = pd.read_excel('Hurricane_Maria_Report_new.xlsx',skiprows = [0,1,2],
                             parse_dates = ['Date Signed','Est. Ultimate Completion Date'])
    df_Harvey = pd.read_excel('Hurricane_Harvey_Report_new.xlsx', skiprows = [0,1,2],
                              parse_dates = ['Date Signed','Est. Ultimate Completion Date'])
    df_Irma = pd.read_excel('Hurricane_Irma_Report_new.xlsx', skiprows = [0,1,2],
                            parse_dates = ['Date Signed','Est. Ultimate Completion Date'])
    df_Irene = pd.read_excel('Hurricane_Irene_Report_new.xlsx', skiprows = [0,1,2],
                              parse_dates = ['Date Signed','Est. Ultimate Completion Date'])
    
    #cleaning the dataframe: removing certain columns and rows
    print("For Hurricane_Maria_Report:")
    df_Maria = Cleaning.IrrCols(df_Maria)
    df_Maria = Cleaning.privateColumns(df_Maria)
    df_Maria = Cleaning.zeroDollars(df_Maria)

    print("For Hurricane_Harvey_Report:")
    df_Harvey = Cleaning.IrrCols(df_Harvey)
    df_Harvey = Cleaning.privateColumns(df_Harvey)
    df_Harvey = Cleaning.zeroDollars(df_Harvey)

    print("For Hurricane_Irma_Report:")
    df_Irma = Cleaning.IrrCols(df_Irma)
    df_Irma = Cleaning.privateColumns(df_Irma)
    df_Irma = Cleaning.zeroDollars(df_Irma)
    
    print("For Hurricane_Irene_Report:")
    df_Irene = Cleaning.IrrCols(df_Irene)
    df_Irene = Cleaning.privateColumns(df_Irene)
    df_Irene = Cleaning.zeroDollars(df_Irene)
    
    #Merging the 3 dataframe to form a single dataset csv file
    final_Dataframe = OneDataFrame.mergeData(df_Maria,df_Harvey,df_Irma,df_Irene)
    final_Dataframe.to_csv('FPDS_final.csv')
    
    
    
    
if __name__ == "__main__":
    main()
        
