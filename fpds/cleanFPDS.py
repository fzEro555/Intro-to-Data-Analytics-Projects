# -*- coding: utf-8 -*-
"""
Created on Wed Oct 31 17:42:18 2018

@author: subha
"""

import numpy as np
import pandas as pd
import openpyxl


# keeping only relevant columns
def cleanup(myDataframe):
    myDataframe.dropna()
    realData = myDataframe[['Contracting Agency Name','Date Signed',
                            'Est. Ultimate Completion Date',
                            'Product or Service Description',
                            'Vendor Address State Name',
                            'National Interest Action','Action Obligation']]
    return realData


# keeping data instances only if action obligation !=0
def zeroDollars(myDataframe):
    valuelist = ["0.00"]
    myDataframe = myDataframe[~myDataframe["Action Obligation"].isin(valuelist)]

    return myDataframe[:-1]


# removes the summary page
def removeSummary(workbook):
    sheetnames = workbook.get_sheet_names()
    std = workbook.get_sheet_by_name(sheetnames[0])
    workbook.remove_sheet(std)
    return workbook


# merges all dataframes into 1
def mergeData(dataFrame1,dataFrame2,dataFrame3,dataFrame4):
    frames = [dataFrame1,dataFrame2,dataFrame3,dataFrame4]
    result = pd.concat(frames)
    return result


def makeData(myDataframe):
    myDataframe = cleanup(myDataframe)
    myDataframe = zeroDollars(myDataframe)
    return myDataframe


if __name__ == "__main__":
    wb = openpyxl.load_workbook('Hurricane_Maria_Report.xlsx')
    workbook_Maria = removeSummary(wb)
    workbook_Maria.save('Hurricane_Maria_Report_new.xlsx')
    wb = openpyxl.load_workbook('Hurricane_Harvey_Report.xlsx')
    workbook_Harvey = removeSummary(wb)
    workbook_Harvey.save('Hurricane_Harvey_Report_new.xlsx')
    wb = openpyxl.load_workbook('Hurricane_Irma_Report.xlsx')
    workbook_Irma = removeSummary(wb)
    workbook_Irma.save('Hurricane_Irma_Report_new.xlsx')
    wb = openpyxl.load_workbook('Hurricane_Irene_Report.xlsx')
    workbook_Irene = removeSummary(wb)
    workbook_Irene.save('Hurricane_Irene_Report_new.xlsx')
    
    df_maria = pd.read_excel('Hurricane_Maria_Report_new.xlsx',skiprows = [0,1,2],
                             parse_dates = ['Date Signed','Est. Ultimate Completion Date'])
    df_maria = makeData(df_maria)
    df_harvey = pd.read_excel('Hurricane_Harvey_Report_new.xlsx', skiprows = [0,1,2],
                              parse_dates = ['Date Signed','Est. Ultimate Completion Date'])
    df_harvey = makeData(df_harvey)
    df_irma = pd.read_excel('Hurricane_Irma_Report_new.xlsx', skiprows = [0,1,2],
                            parse_dates = ['Date Signed','Est. Ultimate Completion Date'])
    df_irma = makeData(df_irma)
    df_irene = pd.read_excel('Hurricane_Irene_Report_new.xlsx', skiprows = [0,1,2],
                              parse_dates = ['Date Signed','Est. Ultimate Completion Date'])
    df_irene = makeData(df_irene)
    final_dataframe = mergeData(df_maria,df_harvey,df_irene,df_irma)
    final_dataframe.to_csv('FPDS_final.csv')

